import requests
import time
import random
import json
from io import BytesIO
from pathlib import Path

import openpyxl


class ONS():

    def __init__(self):

        self.MAX_RETRIES = 5
        # Latest mid-year population estimate (UK, England, Wales, Scotland,
        # Northern Ireland). ONS releases a new mid-year edition annually, so
        # this URL needs bumping (e.g. mid2025/mye25tablesuk.xlsx) once a new
        # edition is published.
        self.XLSX_URL = (
            "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/"
            "populationandmigration/populationestimates/datasets/"
            "populationestimatesforukenglandandwalesscotlandandnorthernireland/"
            "mid2024/mye24tablesuk.xlsx"
        )
        self.HEADERS = {
            "User-Agent": "peptides-uk-interest-analysis/1.0"
        }
        # The workbook bundles 14 sheets (single-year-of-age breakdowns,
        # components of change, density, median age, etc). Only "MYE4" is the
        # actual population count summary, and it holds every geography level
        # (country, region, local authority) in one table, keyed by the
        # "Geography" column.
        self.SUMMARY_SHEET = "MYE4"
        self.NATION_CODES = {
            "K02000001": "United Kingdom",
            "E92000001": "England",
            "S92000003": "Scotland",
            "W92000004": "Wales",
            "N92000002": "Northern Ireland",
        }
        # The 9 English regions plus Wales/Scotland/Northern Ireland: the
        # same 12-way breakdown used for "which region is interest growing
        # in", now carrying actual population counts rather than just names.
        self.REGION_CODES = {
            "E12000001": "North East",
            "E12000002": "North West",
            "E12000003": "Yorkshire and The Humber",
            "E12000004": "East Midlands",
            "E12000005": "West Midlands",
            "E12000006": "East",
            "E12000007": "London",
            "E12000008": "South East",
            "E12000009": "South West",
            "W92000004": "Wales",
            "S92000003": "Scotland",
            "N92000002": "Northern Ireland",
        }

    def fetch_population_estimates(self):

        for attempt in range(self.MAX_RETRIES):

            response = requests.get(
                self.XLSX_URL,
                headers=self.HEADERS,
                timeout=30
            )

            if response.status_code != 429:
                response.raise_for_status()
                return response.content

            wait_seconds = 60 * (2 ** attempt)
            wait_seconds += random.uniform(0, 5)

            print(
                f"Waiting {wait_seconds:.1f} seconds "
                f"before retrying..."
            )

            time.sleep(wait_seconds)

        raise RuntimeError(
            f"ONS continued returning HTTP 429 "
            f"after {self.MAX_RETRIES} attempts"
        )

    def _read_summary_sheet(self, xlsx_bytes):

        workbook = openpyxl.load_workbook(
            BytesIO(xlsx_bytes),
            data_only=True
        )
        sheet = workbook[self.SUMMARY_SHEET]

        header_row = None
        for row in sheet.iter_rows(min_row=1, max_row=20):
            if row[0].value == "Code" and row[1].value == "Name":
                header_row = row[0].row
                break

        year_columns = {}
        for cell in sheet[header_row][3:]:
            if isinstance(cell.value, str) and cell.value.startswith("Mid-"):
                year_columns[cell.value.replace("Mid-", "")] = cell.column

        # single pass over every row (country, region, local authority) so
        # both the nation-level and region-level extracts can reuse it,
        # instead of rescanning the ~400-row sheet once per target code.
        population_by_code = {}
        for row in sheet.iter_rows(min_row=header_row + 1):
            code = row[0].value
            if code is None:
                continue
            population_by_code[code] = {
                year: sheet.cell(row=row[0].row, column=col).value
                for year, col in year_columns.items()
            }

        return population_by_code

    def extract_summary(self, xlsx_bytes, codes, source):

        population_by_code = self._read_summary_sheet(xlsx_bytes)

        geographies = [
            {
                "code": code,
                "name": name,
                "population_by_year": population_by_code[code],
            }
            for code, name in codes.items()
        ]

        return {
            "source": source,
            "unit": "persons",
            "geographies": geographies,
        }

    def save_raw(self, data, filename):

        output = Path(f"data/raw/ons/{filename}")
        output.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        with open(output, "wb") as f:
            f.write(data)

        return data

    def save_json(self, data, filename):

        output = Path(f"data/raw/ons/{filename}")
        output.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        with open(output, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)

        return data


ons = ONS()

if __name__ == "__main__":

    try:

        xlsx_bytes = ons.fetch_population_estimates()

        ons.save_raw(
            xlsx_bytes,
            "population_estimates_uk_source.xlsx"
        )

        nation_summary = ons.extract_summary(
            xlsx_bytes,
            ons.NATION_CODES,
            "ONS MYE4 - Population estimates: Summary for the UK, "
            "mid-2011 to mid-2024"
        )
        ons.save_json(
            nation_summary,
            "population_estimates_uk.json"
        )

        region_summary = ons.extract_summary(
            xlsx_bytes,
            ons.REGION_CODES,
            "ONS MYE4 - Population estimates: Summary for the UK, "
            "mid-2011 to mid-2024 (English regions + Wales/Scotland/"
            "Northern Ireland)"
        )
        ons.save_json(
            region_summary,
            "regions_uk.json"
        )

        print("ONS population estimates retrieved")

    except requests.RequestException as error:
        print(
            f"requests.RequestException: {error}"
        )
    finally:
        time.sleep(6)
